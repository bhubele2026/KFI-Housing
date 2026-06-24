#!/usr/bin/env python3
"""Rent backfill proposal (read-only harvest, NO prod writes).

Proposes monthlyRent for the live properties with monthlyRent=0, pulling ONLY
sourced figures (never guessing). Sources, by confidence:
  - lease-fixes-from-sources.json  (confirmed per-unit lease rents -> summed per property)
  - HARVEST-APPLY-STATUS.md         (documented OCR/lease rents applied 2026-06-23)
  - property tab rent cells         (e.g. Schuette "$1,410")
Anything without a source -> proposedMonthlyRent=null + needsReview=true.
Also emits the per-client weekly-cost table from the master 'Address' tab.
"""
import json, re, urllib.request
from pathlib import Path
import openpyxl

HERE = Path(__file__).parent
MASTER = "/Users/bhubele/Downloads/Housing Master File 2026.xlsx"

def norm(s):
    s = re.sub(r'[^a-z0-9]+', ' ', str(s or '').lower()).strip()
    return s

# --- live zero-rent properties ---
live = json.loads(urllib.request.urlopen("https://kfi-housing.replit.app/api/properties").read())
live = live if isinstance(live, list) else live.get("items", [])
zero = [p for p in live if (p.get("monthlyRent") or 0) == 0]

# --- source 1: lease-fixes (sum per propertyId) ---
rent_by_pid = {}
src_by_pid = {}
try:
    lf = json.load(open(HERE / "lease-fixes-from-sources.json"))
    by = {}
    for x in lf:
        pid = x.get("propertyId"); mr = x.get("monthlyRent")
        if pid and isinstance(mr, (int, float)) and mr > 0:
            by.setdefault(pid, 0); by[pid] += mr
    for pid, tot in by.items():
        rent_by_pid[pid] = tot; src_by_pid[pid] = "lease-fixes-from-sources.json (sum of unit rents)"
except Exception as e:
    print("lease-fixes read failed:", e)

# --- source 2: HARVEST-APPLY-STATUS.md documented rents (name -> rent) ---
# Parse lines like "rent $1,299" / "$7,800/mo" / "$27,840/mo flat" near a property name.
HARVEST_RENTS = []  # (name_substr, monthlyRent, note)
try:
    txt = (HERE / "HARVEST-APPLY-STATUS.md").read_text()
    # known documented monthly rents from the 2026-06-23 apply
    for pat, name, note in [
        (r"Sunset Place.*?\$1,299", "sunset-place", "signed PDF Unit 134"),
        (r"Ridge.*?\$27,840", "1779300519785", "merged Ridge record $27,840/mo"),
        (r"TB Rentals.*?\$7,800", "tb-rentals", "OCR MO commercial lease"),
    ]:
        if re.search(pat, txt, re.S):
            m = re.search(r"\$([\d,]+)", re.search(pat, txt, re.S).group(0))
            if m: HARVEST_RENTS.append((name, int(m.group(1).replace(",", "")), note))
except Exception as e:
    print("harvest-status read failed:", e)

# --- source 3: property-tab rent cells (Schuette $1,410) ---
TAB_RENT = []  # (name_substr, monthlyRent, note)
wb = openpyxl.load_workbook(MASTER, data_only=True)
sch = wb["Schuette.Wausau,WI"]
for r in range(1, sch.max_row + 1):
    for c in range(1, sch.max_column + 1):
        v = sch.cell(r, c).value
        if isinstance(v, str):
            m = re.search(r"\$([\d,]{3,})", v)
            if m:
                TAB_RENT.append(("schuette", int(m.group(1).replace(",", "")), f"tab cell '{v.strip()[:30]}'"))
                break
# dedupe schuette to one value
sch_rent = TAB_RENT[0][1] if TAB_RENT else None

# --- match each zero-rent property ---
def find_rent(p):
    pid = p["id"]; nm = norm(p.get("name"));
    if pid in rent_by_pid:
        return rent_by_pid[pid], src_by_pid[pid], "high"
    for sub, rent, note in HARVEST_RENTS:
        if sub in pid:
            return rent, f"HARVEST-APPLY-STATUS.md: {note}", "high"
    if "schuette" in pid and sch_rent:
        return sch_rent, f"Schuette tab: {TAB_RENT[0][2]}", "high"
    return None, "", ""

proposals = []
for p in zero:
    rent, source, conf = find_rent(p)
    proposals.append({
        "propertyId": p["id"], "name": p.get("name"), "city": p.get("city"),
        "state": p.get("state"), "status": p.get("status"),
        "proposedMonthlyRent": rent, "source": source or "no source in local files",
        "confidence": conf or "none", "needsReview": rent is None,
    })

# --- per-client weekly cost from Address tab ---
addr = wb["Address"]
weekly = []
for r in range(1, addr.max_row + 1):
    cust = addr.cell(r, 1).value; wk = addr.cell(r, 3).value
    if isinstance(cust, str) and cust not in ("Customer",) and isinstance(wk, (int, float)):
        weekly.append((cust.strip(), wk))

withrent = [x for x in proposals if x["proposedMonthlyRent"] is not None]
needs = [x for x in proposals if x["needsReview"]]

# --- write proposal ---
out = ["# Rent backfill proposal (review before applying — NO prod writes made)\n",
       f"- Zero-rent live properties: **{len(zero)}**",
       f"- Confident proposed rent (sourced): **{len(withrent)}**",
       f"- needsReview (no rent in local sources — get from lease PDF / Bill.com / SharePoint Lease Summary): **{len(needs)}**\n",
       "## ✅ Proposed (sourced)"]
for x in sorted(withrent, key=lambda z: -(z["proposedMonthlyRent"] or 0)):
    out.append(f"- **{x['name']}** (`{x['propertyId']}`) → **${x['proposedMonthlyRent']:,}/mo** · {x['source']} [{x['confidence']}]")
out.append("\n## ⚠️ needsReview (no local source — left for lease/Bill.com/SharePoint)")
for x in needs:
    out.append(f"- {x['name']} (`{x['propertyId']}`) · {x['city']} {x['state']} · {x['status']}")
out.append("\n## Per-client weekly cost to associate (master 'Address' tab)")
for cust, wk in weekly:
    out.append(f"- {cust}: ${wk}/wk")
(HERE / "rent-backfill-proposal.md").write_text("\n".join(out))
json.dump(proposals, open(HERE / "rent-backfill-proposal.json", "w"), indent=1)

print(f"zero-rent={len(zero)} | sourced rent={len(withrent)} | needsReview={len(needs)}")
print("PROPOSED:")
for x in withrent: print(f"  {x['name'][:34]:36} ${x['proposedMonthlyRent']:,}/mo  ({x['confidence']})")
print(f"\nweekly-cost rows: {len(weekly)}")
print("proposal: imports/rent-backfill-proposal.md (+ .json)")
